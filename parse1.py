import docx
from docx import Document
from openpyxl import Workbook


def get_input_xls_file_content(source_file):
    if not source_file:
        source_file = '/opt/techgig/MLTickets/docs/SampleInputDoc1-FAQs.docx'
    doc_obj = Document(source_file)
    paras = doc_obj.paragraphs
    return paras

def get_worksheet_obj():
    wb = Workbook()
    ws = wb.create_sheet("hardware")
    ws.title = "hardware_resolution"
    return [wb, ws]

def write_to_xls(source_file, dest_file):
    work_book, work_sheet = get_worksheet_obj()
    res_dict = {}
    title = None
    resolution = None
    row_count = 1
    paras = get_input_xls_file_content(source_file)
    #for row in range(1,len(paras)-50):
    title=None
    resolutions = []
    qna = {}
    for row in range(0,len(paras)):
        para = paras[row]
        for run in para.runs:
            if run.bold:
                if title:
                    qna[title] = resolutions
                    resolutions = []
                title = run.text
            else:
                resolutions.append(run.text)

    for i in qna.keys():            
        for j in qna[i]:
            work_sheet.cell(column=1, row=row_count, value=i)
            work_sheet.cell(column=2, row=row_count, value=j)
            row_count += 1
            
    work_book.save(dest_file)

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Please provide input docs file to convert into xlsx file")
    else:
        source_file = sys.argv[1]
        if sys.argv[2]:
            destination_file = sys.argv[2].strip()
        else:
            destination_file = None
        get_worksheet_obj()
        import pudb;pu.db
        write_to_xls(source_file, destination_file)
